import json
import openpyxl
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.http import JsonResponse
from django.core.paginator import Paginator
from django.views.decorators.http import require_POST
from django.forms import formset_factory
from .models import TableStructure, TableColumn, TableData, TableCell, ExcelImport
from .forms import TableStructureForm, TableColumnForm, ExcelImportForm, HeaderMappingForm, TableDataForm


def index(request):
    """Main page with tabs."""
    tables = TableStructure.objects.all()
    return render(request, 'table_app/index.html', {'tables': tables})


def create_table(request):
    """View for creating a new table."""
    if request.method == 'POST':
        form = TableStructureForm(request.POST)
        if form.is_valid():
            table = form.save()
            return redirect('add_columns', table_id=table.id)
        
    else:
        form = TableStructureForm()
    
    return render(request, 'table_app/create_table.html', {'form': form})


def add_columns(request, table_id):
    """View for adding columns to a table."""
    table = get_object_or_404(TableStructure, id=table_id)
    columns = table.columns.all()
    
    if request.method == 'POST':
        form = TableColumnForm(request.POST)
        if form.is_valid():
            column = form.save(commit=False)
            column.table = table
            column.order = columns.count()  # Set order to the current count
            column.save()
            return redirect('add_columns', table_id=table.id)
    else:
        form = TableColumnForm()
    
    return render(request, 'table_app/add_columns.html', {
        'table': table,
        'columns': columns,
        'form': form
    })


def delete_column(request, column_id):
    """Delete a column."""
    column = get_object_or_404(TableColumn, id=column_id)
    table_id = column.table.id
    column.delete()
    return redirect('add_columns', table_id=table_id)


def manage_data(request, table_id):
    """View for managing data (add data or import from Excel)."""
    table = get_object_or_404(TableStructure, id=table_id)
    
    # Form for adding data manually
    if request.method == 'POST' and 'add_data' in request.POST:
        data_form = TableDataForm(request.POST, table_structure=table)
        if data_form.is_valid():
            # Create a new data row
            data_row = TableData.objects.create(table=table)
            
            # Create cells for each column
            for column in table.columns.all():
                field_name = f"column_{column.id}"
                value = data_form.cleaned_data.get(field_name)
                
                cell = TableCell(row=data_row, column=column)
                cell.set_value(value)
                cell.save()
                
            return redirect('manage_data', table_id=table.id)
    else:
        data_form = TableDataForm(table_structure=table)
    
    # Form for importing from Excel
    if request.method == 'POST' and 'import_excel' in request.POST:
        excel_form = ExcelImportForm(request.POST, request.FILES)
        if excel_form.is_valid():
            import_obj = excel_form.save(commit=False)
            import_obj.table = table
            import_obj.save()
            
            # Extract headers from Excel
            wb = openpyxl.load_workbook(import_obj.file.path)
            ws = wb.active
            excel_headers = [str(cell.value) for cell in ws[1] if cell.value]
            
            # Get table column names
            table_columns = list(table.columns.all().values('id', 'name'))
            
            return render(request, 'table_app/match_headers.html', {
                'table': table,
                'excel_import': import_obj,
                'excel_headers': excel_headers,
                'table_columns': table_columns
            })
    else:
        excel_form = ExcelImportForm()
    
    return render(request, 'table_app/manage_data.html', {
        'table': table,
        'data_form': data_form,
        'excel_form': excel_form
    })


@require_POST
def save_header_mapping(request, import_id):
    """Save the header mapping and import data."""
    excel_import = get_object_or_404(ExcelImport, id=import_id)
    table = excel_import.table
    
    mapping = json.loads(request.POST.get('mapping', '{}'))
    excel_import.set_header_mapping(mapping)
    excel_import.save()
    
    # Load Excel data
    wb = openpyxl.load_workbook(excel_import.file.path)
    ws = wb.active
    
    # Skip header row
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    
    # Process each row
    for row_values in rows:
        # Create a new data row
        data_row = TableData.objects.create(table=table)
        
        # Map Excel columns to table columns and create cells
        for excel_header, column_id in mapping.items():
            if not column_id:  # Skip unmapped headers
                continue
                
            try:
                column = TableColumn.objects.get(id=column_id)
                
                # Find the index of this header in the Excel file
                excel_headers = [str(cell.value) for cell in ws[1] if cell.value]
                if excel_header in excel_headers:
                    excel_index = excel_headers.index(excel_header)
                    value = row_values[excel_index] if excel_index < len(row_values) else None
                    
                    cell = TableCell(row=data_row, column=column)
                    cell.set_value(value)
                    cell.save()
            except (TableColumn.DoesNotExist, ValueError, IndexError):
                continue
    
    return redirect('view_data', table_id=table.id)


def view_data(request, table_id):
    """View table data with pagination."""
    table = get_object_or_404(TableStructure, id=table_id)
    columns = table.columns.all()
    data_rows = TableData.objects.filter(table=table).order_by('-created_at')
    
    # Prepare data for display
    rows = []
    for data_row in data_rows:
        row_data = {'id': data_row.id, 'cells': {}}
        for column in columns:
            try:
                cell = TableCell.objects.get(row=data_row, column=column)
                row_data['cells'][column.id] = cell.get_value()
            except TableCell.DoesNotExist:
                row_data['cells'][column.id] = None
        rows.append(row_data)
    
    # Pagination
    paginator = Paginator(rows, 10)  # Show 10 rows per page
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'table_app/view_data.html', {
        'table': table,
        'columns': columns,
        'page_obj': page_obj
    })


def delete_table(request, table_id):
    """Delete a table."""
    table = get_object_or_404(TableStructure, id=table_id)
    if request.method == 'POST':
        table.delete()
        return redirect('index')
    return render(request, 'table_app/delete_table.html', {'table': table})